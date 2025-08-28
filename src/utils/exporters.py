import json
import logging
import time
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from config.settings import (
    CSV_SEPARATOR,
    DEFAULT_CSV_FILENAME,
    DEFAULT_EXCEL_FILENAME,
    DEFAULT_JSON_FILENAME,
    EXPORTS_DIR,
)

from ..core.models import Vacancy

logger = logging.getLogger(__name__)


class ExcelExporter:
    def export_to_excel(self, vacancies: List[Vacancy], filename: Optional[str] = None) -> str:
        """Экспорт вакансий в Excel с правильной кодировкой и форматированием"""
        if not vacancies:
            return "Нет данных для экспорта"

        filename = filename or DEFAULT_EXCEL_FILENAME
        logger.info(f"Экспорт {len(vacancies)} вакансий в Excel")
        start_time = time.time()

        try:
            filepath = EXPORTS_DIR / filename

            # Создаем новую книгу Excel
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Вакансии"

                # Подготавливаем данные
                data = self._prepare_export_data(vacancies)

                # Заголовки
                headers = list(data[0].keys()) if data else []
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

                # Данные
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, (key, value) in enumerate(row_data.items(), 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Автоподбор ширины колонок
                for column in ws.columns:
                    max_length = 0
                    first_cell = column[0]
                    if hasattr(first_cell, "column_letter") and first_cell.column_letter:
                        column_letter = first_cell.column_letter
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            wb.save(filepath)

            elapsed = time.time() - start_time
            logger.info(f"Excel экспорт завершен за {elapsed:.2f} сек")
            return str(filepath)

        except Exception as e:
            logger.error(f"Ошибка Excel экспорта: {e}")
            return f"Ошибка: {e}"

    def _prepare_export_data(self, vacancies: List[Vacancy]) -> List[Dict[str, Any]]:
        """Подготовка данных для экспорта"""
        data = []
        for vacancy in vacancies:
            salary_from = vacancy.salary.from_amount if vacancy.salary else None
            salary_to = vacancy.salary.to_amount if vacancy.salary else None
            currency = vacancy.salary.currency if vacancy.salary else None

            # Обрабатываем текст для корректного отображения
            snippet = self._clean_text(vacancy.snippet) if vacancy.snippet else ""
            experience = self._clean_text(vacancy.experience) if vacancy.experience else ""
            employment = self._clean_text(vacancy.employment) if vacancy.employment else ""

            data.append(
                {
                    "ID": vacancy.id,
                    "Название": self._clean_text(vacancy.name),
                    "Компания": self._clean_text(vacancy.company),
                    "Зарплата от": salary_from,
                    "Зарплата до": salary_to,
                    "Валюта": currency,
                    "Город": self._clean_text(vacancy.area),
                    "Опыт": experience,
                    "Тип занятости": employment,
                    "Ссылка": vacancy.url,
                    "Источник": vacancy.source,
                    "Описание": snippet,
                    "Дата публикации": vacancy.published_at,
                }
            )
        return data

    def _clean_text(self, text: str) -> str:
        """Очистка текста от HTML тегов и специальных символов"""
        if not text:
            return ""

        # Удаляем HTML теги
        import re

        clean_text = re.sub(r"<[^>]+>", "", text)

        # Заменяем HTML entities
        clean_text = clean_text.replace("&nbsp;", " ").replace("&amp;", "&")

        # Удаляем лишние пробелы
        clean_text = " ".join(clean_text.split())

        return clean_text


class CSVExporter:
    def export_to_csv(self, vacancies: List[Vacancy], filename: Optional[str] = None) -> str:
        """Экспорт в CSV с правильной кодировкой"""
        if not vacancies:
            return "Нет данных для экспорта"

        filename = filename or DEFAULT_CSV_FILENAME
        logger.info(f"Экспорт {len(vacancies)} вакансий в CSV")
        start_time = time.time()

        try:
            data = self._prepare_export_data(vacancies)
            df = pd.DataFrame(data)
            filepath = EXPORTS_DIR / filename

            # Используем encoding='utf-8-sig' для корректного отображения в Excel
            df.to_csv(filepath, index=False, encoding="utf-8-sig", sep=CSV_SEPARATOR)

            elapsed = time.time() - start_time
            logger.info(f"CSV экспорт завершен за {elapsed:.2f} сек")
            return str(filepath)

        except Exception as e:
            logger.error(f"Ошибка CSV экспорта: {e}")
            return f"Ошибка: {e}"

    def _prepare_export_data(self, vacancies: List[Vacancy]) -> List[Dict[str, Any]]:
        """Подготовка данных для экспорта"""
        data = []
        for vacancy in vacancies:
            salary_from = vacancy.salary.from_amount if vacancy.salary else None
            salary_to = vacancy.salary.to_amount if vacancy.salary else None
            currency = vacancy.salary.currency if vacancy.salary else None

            data.append(
                {
                    "ID": vacancy.id,
                    "Название": vacancy.name,
                    "Компания": vacancy.company,
                    "Зарплата от": salary_from,
                    "Зарплата до": salary_to,
                    "Валюта": currency,
                    "Город": vacancy.area,
                    "Опыт": vacancy.experience,
                    "Тип занятости": vacancy.employment,
                    "Ссылка": vacancy.url,
                    "Источник": vacancy.source,
                }
            )
        return data


class JSONExporter:
    def export_to_json(self, vacancies: List[Vacancy], filename: Optional[str] = None) -> str:
        """Экспорт вакансий в JSON"""
        if not vacancies:
            return "Нет данных для экспорта"

        filename = filename or DEFAULT_JSON_FILENAME
        logger.info(f"Экспорт {len(vacancies)} вакансий в JSON")
        start_time = time.time()

        try:
            data = [vacancy.to_dict() for vacancy in vacancies]
            filepath = EXPORTS_DIR / filename

            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            elapsed = time.time() - start_time
            logger.info(f"JSON экспорт завершен за {elapsed:.2f} сек")
            return str(filepath)

        except Exception as e:
            logger.error(f"Ошибка JSON экспорта: {e}")
            return f"Ошибка: {e}"
